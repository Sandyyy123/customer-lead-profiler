"""
Builds a formatted Excel dashboard from segmentation + comparison outputs.
Produces a multi-sheet workbook with PivotTable-ready data, charts, and summary.
Requires: openpyxl
"""

import pandas as pd
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


ACCENT_PURPLE = "6C5CE7"
ACCENT_GREEN = "00B894"
ACCENT_ORANGE = "FD9644"
DARK_BG = "12121A"
LIGHT_TEXT = "E2E2F0"
HEADER_FILL = PatternFill("solid", fgColor=ACCENT_PURPLE)
HEADER_FONT = Font(bold=True, color=LIGHT_TEXT)
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


def style_header_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=col_name)
    style_header_row(ws, start_row, len(df.columns))

    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F4FF")

    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].auto_size = True
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            len(str(df.columns[col_idx - 1])) + 4, 14
        )


def build_summary_sheet(ws, segments_df: pd.DataFrame) -> None:
    ws.title = "Summary"
    ws["A1"] = "Customer & Lead Analysis Dashboard"
    ws["A1"].font = Font(bold=True, size=16, color=ACCENT_PURPLE)
    ws["A2"] = "Segments, buyer differences, and geographic opportunities — all in one place"
    ws["A2"].font = Font(italic=True, color="888888", size=11)

    if "segment" in segments_df.columns:
        seg_summary = (
            segments_df.groupby("segment")
            .agg(count=("customer_id", "count"))
            .reset_index()
        )
        ws["A4"] = "Segment Breakdown"
        ws["A4"].font = Font(bold=True, size=13)
        write_df_to_sheet(ws, seg_summary, start_row=5)

        chart = BarChart()
        chart.title = "Customers per Segment"
        chart.style = 10
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Segment"
        n = len(seg_summary)
        data = Reference(ws, min_col=2, min_row=5, max_row=5 + n)
        cats = Reference(ws, min_col=1, min_row=6, max_row=5 + n)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D5")


def build(segments_path: str, comparison_path: str, geo_path: str, output_path: str) -> None:
    wb = Workbook()

    segments_df = pd.read_csv(segments_path) if segments_path else pd.DataFrame()
    comparison_df = pd.read_csv(comparison_path) if comparison_path else pd.DataFrame()
    geo_df = pd.read_csv(geo_path) if geo_path else pd.DataFrame()

    ws_summary = wb.active
    build_summary_sheet(ws_summary, segments_df)

    if not segments_df.empty:
        ws_seg = wb.create_sheet("Customer Segments")
        write_df_to_sheet(ws_seg, segments_df)

    if not comparison_df.empty:
        ws_cmp = wb.create_sheet("Buyer vs Non-Buyer")
        write_df_to_sheet(ws_cmp, comparison_df)

    if not geo_df.empty:
        ws_geo = wb.create_sheet("Geographic Analysis")
        write_df_to_sheet(ws_geo, geo_df)

    wb.save(output_path)
    print(f"Dashboard saved to {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build Excel dashboard from analysis outputs")
    parser.add_argument("--segments", required=True, help="Segmented customers CSV")
    parser.add_argument("--comparison", required=True, help="Buyer vs non-buyer CSV")
    parser.add_argument("--geo", required=True, help="Geographic analysis CSV")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()
    build(args.segments, args.comparison, args.geo, args.output)
